from __future__ import annotations

import csv
import io
import threading
import uuid
from collections import defaultdict, deque
from dataclasses import dataclass, field
from datetime import datetime, timezone
from typing import Dict, Iterable, List, MutableMapping, Optional, Sequence, Set, Tuple

from flask import session

DATA_IMPORT_SESSION_KEY = "data_import_session_id"

DATA_IMPORT_OBJECTS: List[Dict[str, str]] = [
    {"key": "Account", "label": "Account"},
    {"key": "BillingProfile", "label": "Billing Profile"},
    {"key": "Contact", "label": "Contact"},
    {"key": "Contract", "label": "Contract"},
    {"key": "AccountContactRelationship", "label": "Account Contact Relationship"},
    {"key": "Individual", "label": "Individual"},
    {"key": "ContactPointPhone", "label": "Contact Point Phone"},
    {"key": "ContactPointEmail", "label": "Contact Point Email"},
    {"key": "Case", "label": "Case"},
    {"key": "Order", "label": "Order"},
    {"key": "Sale", "label": "Sale"},
]

_OBJECT_LOOKUP = {item["key"]: item for item in DATA_IMPORT_OBJECTS}


@dataclass
class ObjectDataset:
    fields: List[str]
    records: Dict[str, Dict[str, str]]
    filename: str
    updated_at: datetime = field(default_factory=lambda: datetime.now(timezone.utc))


class ImportSession:
    def __init__(self) -> None:
        self.objects: Dict[str, ObjectDataset] = {}
        self._links: MutableMapping[Tuple[str, str], Set[Tuple[str, str]]] = defaultdict(set)

    def set_object_data(self, object_key: str, dataset: ObjectDataset) -> None:
        self.objects[object_key] = dataset
        self._rebuild_links()

    def clear_object_data(self, object_key: str) -> bool:
        removed = self.objects.pop(object_key, None) is not None
        if removed:
            self._rebuild_links()
        return removed

    def _rebuild_links(self) -> None:
        self._links = defaultdict(set)
        id_lookup: Dict[str, Tuple[str, str]] = {}
        for object_key, dataset in self.objects.items():
            for record_id in dataset.records.keys():
                id_lookup[record_id] = (object_key, record_id)

        for object_key, dataset in self.objects.items():
            for record_id, record in dataset.records.items():
                node = (object_key, record_id)
                self._links.setdefault(node, set())
                for field_name, value in record.items():
                    if not field_name or field_name.lower() == "id":
                        continue
                    for candidate in _extract_candidate_ids(value):
                        target = id_lookup.get(candidate)
                        if not target or target == node:
                            continue
                        self._links[node].add(target)
                        self._links[target].add(node)

    def get_status(self) -> List[Dict[str, object]]:
        status = []
        for definition in DATA_IMPORT_OBJECTS:
            object_key = definition["key"]
            dataset = self.objects.get(object_key)
            if not dataset:
                status.append(
                    {
                        "key": object_key,
                        "label": definition["label"],
                        "loaded": False,
                        "recordCount": 0,
                        "fields": [],
                        "filename": None,
                        "updatedAt": None,
                    }
                )
                continue
            status.append(
                {
                    "key": object_key,
                    "label": definition["label"],
                    "loaded": True,
                    "recordCount": len(dataset.records),
                    "fields": list(dataset.fields),
                    "filename": dataset.filename,
                    "updatedAt": dataset.updated_at.isoformat(),
                }
            )
        return status

    def get_fields(self, object_key: str) -> List[str]:
        dataset = self.objects.get(object_key)
        if not dataset:
            return []
        return list(dataset.fields)

    def get_autocomplete_values(
        self, object_key: str, field_name: str, term: str, limit: int = 12
    ) -> List[str]:
        dataset = self.objects.get(object_key)
        if not dataset or field_name not in dataset.fields:
            return []
        normalized_term = (term or "").strip().lower()
        unique_values: Set[str] = set()
        for record in dataset.records.values():
            value = (record.get(field_name) or "").strip()
            if not value:
                continue
            if normalized_term and normalized_term not in value.lower():
                continue
            unique_values.add(value)
            if len(unique_values) >= limit * 3:
                # Gather extra values for better sorting
                break
        sorted_values = sorted(unique_values)
        if limit:
            return sorted_values[:limit]
        return sorted_values

    def search_records(
        self, object_key: str, field_name: str, value: str
    ) -> List[Tuple[str, str]]:
        dataset = self.objects.get(object_key)
        if not dataset or field_name not in dataset.fields:
            return []
        target = (value or "").strip().lower()
        if not target:
            return []
        matches: List[Tuple[str, str]] = []
        for record_id, record in dataset.records.items():
            candidate = (record.get(field_name) or "").strip().lower()
            if candidate == target:
                matches.append((object_key, record_id))
        return matches

    def get_related_component(
        self, starting_nodes: Sequence[Tuple[str, str]]
    ) -> Dict[str, object]:
        if not starting_nodes:
            return {"objects": {}, "matches": []}
        visited: Set[Tuple[str, str]] = set()
        queue: deque[Tuple[str, str]] = deque(starting_nodes)
        component: Set[Tuple[str, str]] = set()
        while queue:
            node = queue.popleft()
            if node in visited:
                continue
            visited.add(node)
            component.add(node)
            for neighbor in self._links.get(node, set()):
                if neighbor not in visited:
                    queue.append(neighbor)

        objects_payload: Dict[str, Dict[str, object]] = {}
        for object_key, record_id in component:
            dataset = self.objects.get(object_key)
            if not dataset:
                continue
            record = dataset.records.get(record_id)
            if not record:
                continue
            object_info = objects_payload.setdefault(
                object_key,
                {
                    "key": object_key,
                    "label": _OBJECT_LOOKUP.get(object_key, {}).get("label", object_key),
                    "records": [],
                },
            )
            related_nodes = [
                {"object": neighbor_key, "id": neighbor_id}
                for neighbor_key, neighbor_id in sorted(self._links.get((object_key, record_id), set()))
                if (neighbor_key, neighbor_id) in component
            ]
            object_info["records"].append(
                {
                    "id": record_id,
                    "fields": _serialize_record_fields(dataset.fields, record),
                    "related": related_nodes,
                }
            )

        for object_info in objects_payload.values():
            object_info["records"].sort(key=lambda item: item.get("id", ""))

        matches_payload = []
        for object_key, record_id in starting_nodes:
            dataset = self.objects.get(object_key)
            if not dataset:
                continue
            record = dataset.records.get(record_id)
            if not record:
                continue
            matches_payload.append(
                {
                    "object": object_key,
                    "id": record_id,
                    "fields": _serialize_record_fields(dataset.fields, record),
                }
            )

        return {"objects": objects_payload, "matches": matches_payload}


def _serialize_record_fields(fields: Iterable[str], record: Optional[Dict[str, str]]) -> List[Dict[str, str]]:
    record = record or {}
    serialized = []
    for field_name in fields:
        serialized.append({"name": field_name, "value": record.get(field_name, "")})
    return serialized


def _extract_candidate_ids(value: object) -> Iterable[str]:
    if value is None:
        return []
    if isinstance(value, str):
        tokens = []
        for part in value.replace("\n", " ").split(";"):
            cleaned = part.strip()
            if cleaned:
                tokens.append(cleaned)
        return tokens
    return [str(value).strip()]


def _normalize_header(value: object) -> str:
    return str(value).strip()


def _stringify_cell(value: object) -> str:
    if value is None:
        return ""
    if isinstance(value, datetime):
        if value.tzinfo is None:
            value = value.replace(tzinfo=timezone.utc)
        return value.isoformat()
    if isinstance(value, float):
        if value.is_integer():
            return str(int(value))
    return str(value)


def parse_tabular_file(filename: str, file_bytes: bytes) -> ObjectDataset:
    if not filename:
        raise ValueError("missing_filename")
    if not file_bytes:
        raise ValueError("empty_file")
    name_lower = filename.lower()
    if name_lower.endswith(".csv"):
        headers, rows = _parse_csv(file_bytes)
    else:
        headers, rows = _parse_excel(file_bytes)

    normalized_headers = []
    source_headers: Dict[str, str] = {}
    for raw_header in headers:
        normalized = _normalize_header(raw_header)
        if normalized.lower() == "id":
            normalized = "Id"
        normalized_headers.append(normalized)
        source_headers[normalized] = raw_header if isinstance(raw_header, str) else str(raw_header)
    if not normalized_headers or any(not header for header in normalized_headers):
        raise ValueError("invalid_headers")
    lower_headers = [header.lower() for header in normalized_headers]
    if len(set(lower_headers)) != len(lower_headers):
        raise ValueError("duplicate_headers")
    if "id" not in lower_headers:
        raise ValueError("missing_id")

    id_field = next((header for header in normalized_headers if header.lower() == "id"), "Id")
    records: Dict[str, Dict[str, str]] = {}
    for row in rows:
        record: Dict[str, str] = {}
        for field in normalized_headers:
            source_key = source_headers.get(field, field)
            if isinstance(row, dict):
                value = row.get(source_key)
            else:
                value = None
            record[field] = _stringify_cell(value)
        record_id = record.get(id_field, "").strip()
        if not record_id:
            continue
        if record_id not in records:
            records[record_id] = record

    dataset = ObjectDataset(
        fields=normalized_headers,
        records=records,
        filename=filename,
        updated_at=datetime.now(timezone.utc),
    )
    return dataset


def _parse_csv(file_bytes: bytes) -> Tuple[List[str], List[Dict[str, object]]]:
    text_stream = io.StringIO(file_bytes.decode("utf-8-sig"))
    reader = csv.reader(text_stream)
    try:
        headers = next(reader)
    except StopIteration as exc:
        raise ValueError("empty_file") from exc
    rows: List[Dict[str, object]] = []
    for values in reader:
        row = {headers[index]: values[index] if index < len(values) else "" for index in range(len(headers))}
        rows.append(row)
    return headers, rows


def _parse_excel(file_bytes: bytes) -> Tuple[List[str], List[Dict[str, object]]]:
    from openpyxl import load_workbook

    try:
        workbook = load_workbook(io.BytesIO(file_bytes), data_only=True, read_only=True)
    except Exception as exc:  # pragma: no cover - delegated to openpyxl
        raise ValueError("invalid_workbook") from exc
    try:
        sheet = workbook.active
        rows_iter = sheet.iter_rows(values_only=True)
        headers = next(rows_iter, None)
        if not headers:
            raise ValueError("empty_file")
        headers_list = [str(cell) if cell is not None else "" for cell in headers]
        rows: List[Dict[str, object]] = []
        for row_values in rows_iter:
            row_dict: Dict[str, object] = {}
            for index, header in enumerate(headers_list):
                cell_value = row_values[index] if row_values and index < len(row_values) else None
                row_dict[header] = cell_value
            rows.append(row_dict)
        return headers_list, rows
    finally:
        workbook.close()


_sessions: Dict[str, ImportSession] = {}
_sessions_lock = threading.Lock()


def _ensure_session_id() -> str:
    session_id = session.get(DATA_IMPORT_SESSION_KEY)
    if session_id and isinstance(session_id, str):
        return session_id
    session_id = uuid.uuid4().hex
    session[DATA_IMPORT_SESSION_KEY] = session_id
    return session_id


def get_import_session() -> ImportSession:
    session_id = _ensure_session_id()
    with _sessions_lock:
        if session_id not in _sessions:
            _sessions[session_id] = ImportSession()
        return _sessions[session_id]


def get_object_definition(object_key: str) -> Optional[Dict[str, str]]:
    return _OBJECT_LOOKUP.get(object_key)

